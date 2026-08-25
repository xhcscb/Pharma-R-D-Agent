import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from pharma_data.contracts import DocumentType, ElementType, ParsedDocument
from pharma_data.parsers.common import make_element


class MarketDataParser:
    """把授权行情导出文件规范化为逐证券、逐交易日证据元素。"""

    name = "market-data-structured"
    version = "0.1.0"
    field_aliases = {
        "company": ("company", "company_name", "公司", "证券简称"),
        "stock_code": ("stock_code", "code", "证券代码"),
        "trade_date": ("trade_date", "date", "交易日期"),
        "open": ("open", "开盘价"),
        "high": ("high", "最高价"),
        "low": ("low", "最低价"),
        "close": ("close", "收盘价"),
        "volume": ("volume", "成交量"),
        "turnover": ("turnover", "amount", "成交额"),
        "market_cap": ("market_cap", "market cap", "总市值"),
        "currency": ("currency", "币种"),
        "adjustment": ("adjustment", "复权口径"),
    }

    def parse(
        self,
        path: Path,
        *,
        document_id: str,
        document_version_id: str,
        document_type: DocumentType,
        artifact_id: str,
    ) -> ParsedDocument:
        rows = self._rows(path)
        elements = []
        for order, raw_row in enumerate(rows):
            row = self._normalize_row(raw_row)
            elements.append(
                make_element(
                    document_version_id=document_version_id,
                    element_type=ElementType.STRUCTURED_RECORD,
                    reading_order=order,
                    text=self._evidence_text(row),
                    parser_name=self.name,
                    parser_version=self.version,
                    structured_payload={"market_record": row},
                )
            )
        return ParsedDocument(
            document_id=document_id,
            document_version_id=document_version_id,
            document_type=document_type,
            metadata={"artifact_id": artifact_id},
            elements=elements,
            parse_quality={"market_record_count": float(len(elements))},
        )

    def _rows(self, path: Path) -> list[dict[str, Any]]:
        suffix = path.suffix.casefold()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                return [dict(row) for row in csv.DictReader(stream)]
        if suffix == ".json":
            payload = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(payload, dict) and "records" in payload:
                payload = payload["records"]
            if isinstance(payload, dict):
                payload = [payload]
            if not isinstance(payload, list) or not all(isinstance(item, dict) for item in payload):
                raise ValueError("行情 JSON 必须是对象、对象数组或包含 records 的对象")
            return [dict(item) for item in payload]
        if suffix == ".xlsx":
            workbook = load_workbook(path, data_only=True, read_only=True)
            rows: list[dict[str, Any]] = []
            for sheet in workbook.worksheets:
                values = list(sheet.iter_rows(values_only=True))
                if not values:
                    continue
                headers = [str(value or "").strip() for value in values[0]]
                rows.extend(
                    {headers[index]: value for index, value in enumerate(row) if headers[index]}
                    for row in values[1:]
                    if any(value is not None for value in row)
                )
            workbook.close()
            return rows
        raise ValueError("行情文件只支持 JSON、CSV 或 XLSX")

    def _normalize_row(self, raw_row: dict[str, Any]) -> dict[str, Any]:
        folded = {str(key).strip().casefold(): value for key, value in raw_row.items()}
        normalized = {}
        for canonical, aliases in self.field_aliases.items():
            normalized[canonical] = next(
                (folded[alias.casefold()] for alias in aliases if alias.casefold() in folded),
                None,
            )
        return normalized

    @staticmethod
    def _evidence_text(row: dict[str, Any]) -> str:
        company = row.get("company") or "未提供公司名称"
        code = row.get("stock_code") or "未提供证券代码"
        trade_date = row.get("trade_date") or "未提供交易日"
        currency = row.get("currency") or "CNY"
        parts = [f"{company}（{code}）交易日 {trade_date}"]
        fields = (
            ("open", "开盘价", "元"),
            ("high", "最高价", "元"),
            ("low", "最低价", "元"),
            ("close", "收盘价", "元"),
            ("volume", "成交量", "股"),
            ("turnover", "成交额", "元"),
            ("market_cap", "总市值", "元"),
        )
        parts.extend(
            f"{label} {row[key]} {unit}"
            for key, label, unit in fields
            if row.get(key) not in (None, "")
        )
        parts.append(f"币种 {currency}")
        if row.get("adjustment"):
            parts.append(f"复权口径 {row['adjustment']}")
        return "；".join(parts)
