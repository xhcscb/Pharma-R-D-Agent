import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from lxml import etree  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from pharma_data.contracts import DocumentType, ElementType, ParsedDocument
from pharma_data.parsers.base import Parser
from pharma_data.parsers.common import make_element
from pharma_data.utils.hashing import stable_uuid


class JsonParser(Parser):
    name = "json-structure"
    version = "0.2.0"
    media_types = {"application/json"}

    def parse(
        self,
        path: Path,
        *,
        document_id: str,
        document_version_id: str,
        document_type: DocumentType,
        artifact_id: str,
    ) -> ParsedDocument:
        payload = json.loads(path.read_text(encoding="utf-8"))
        elements = [
            make_element(
                document_version_id=document_version_id,
                element_type=ElementType.STRUCTURED_RECORD,
                reading_order=0,
                text="",
                parser_name=self.name,
                parser_version=self.version,
                structured_payload={"record": payload},
            )
        ]
        for order, (json_path, value) in enumerate(self._text_fields(payload), start=1):
            if isinstance(value, dict):
                display_value = json.dumps(value, ensure_ascii=False, sort_keys=True)
            elif isinstance(value, list):
                display_value = "; ".join(str(item) for item in value)
            else:
                display_value = str(value)
            elements.append(
                make_element(
                    document_version_id=document_version_id,
                    element_type=ElementType.PARAGRAPH,
                    reading_order=order,
                    text=f"{json_path}: {display_value}",
                    parser_name=self.name,
                    parser_version=self.version,
                    structured_payload={"json_path": json_path, "value": value},
                )
            )
        return ParsedDocument(
            document_id=document_id,
            document_version_id=document_version_id,
            document_type=document_type,
            metadata={"artifact_id": artifact_id},
            elements=elements,
            parse_quality={"structured": 1.0, "field_elements": float(len(elements) - 1)},
        )

    @classmethod
    def _text_fields(cls, value: Any, path: str = "$") -> list[tuple[str, Any]]:
        fields: list[tuple[str, Any]] = []
        if isinstance(value, dict):
            if "val" in value and ("filed" in value or "end" in value):
                return [(path, value)]
            for key, child in value.items():
                fields.extend(cls._text_fields(child, f"{path}.{key}"))
        elif isinstance(value, list):
            if value and all(not isinstance(item, (dict, list)) for item in value):
                fields.append((path, value))
            else:
                for index, child in enumerate(value):
                    fields.extend(cls._text_fields(child, f"{path}[{index}]"))
        elif value is not None:
            fields.append((path, value))
        return fields


class SpreadsheetParser(Parser):
    name = "openpyxl-financial"
    version = "0.1.0"
    media_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }

    def parse(
        self,
        path: Path,
        *,
        document_id: str,
        document_version_id: str,
        document_type: DocumentType,
        artifact_id: str,
    ) -> ParsedDocument:
        workbook = load_workbook(path, data_only=False, read_only=False)
        elements = []
        order = 0
        for worksheet in workbook.worksheets:
            rows: list[list[Any]] = []
            for row in worksheet.iter_rows(values_only=True):
                values = [cell for cell in row]
                if any(value is not None for value in values):
                    rows.append(values)
            text = "\n".join(
                "\t".join("" if value is None else str(value) for value in row) for row in rows
            )
            merged_cells = [str(item) for item in worksheet.merged_cells.ranges]
            statement_type = self._statement_type(worksheet.title)
            financial_cells = self._financial_cells(
                rows,
                worksheet.title,
                statement_type,
                document_version_id,
            )
            elements.append(
                make_element(
                    document_version_id=document_version_id,
                    element_type=ElementType.TABLE,
                    reading_order=order,
                    text=text,
                    parser_name=self.name,
                    parser_version=self.version,
                    structured_payload={
                        "sheet": worksheet.title,
                        "rows": rows,
                        "merged_cells": merged_cells,
                        "header_levels": 1 if rows else 0,
                        "financial_cells": financial_cells,
                        "financial_fields": {
                            "statement_type": statement_type,
                            "unit": None,
                            "currency": None,
                            "period_start": None,
                            "period_end": None,
                            "consolidation_scope": None,
                            "audited": None,
                            "restated": None,
                        },
                    },
                )
            )
            order += 1
        workbook.close()
        return ParsedDocument(
            document_id=document_id,
            document_version_id=document_version_id,
            document_type=document_type,
            metadata={"artifact_id": artifact_id, "sheet_names": workbook.sheetnames},
            elements=elements,
            parse_quality={
                "sheet_count": float(len(elements)),
                "financial_cell_count": float(
                    sum(
                        len(item.structured_payload.get("financial_cells", [])) for item in elements
                    )
                ),
            },
        )

    @staticmethod
    def _statement_type(sheet_name: str) -> str | None:
        folded = sheet_name.casefold()
        mapping = {
            "income_statement": (
                "income",
                "profit",
                "\u5229\u6da6",
                "\u635f\u76ca",
            ),
            "balance_sheet": ("balance", "\u8d44\u4ea7\u8d1f\u503a"),
            "cash_flow_statement": ("cash", "\u73b0\u91d1\u6d41"),
        }
        return next(
            (
                statement_type
                for statement_type, aliases in mapping.items()
                if any(alias in folded for alias in aliases)
            ),
            None,
        )

    @staticmethod
    def _financial_cells(
        rows: list[list[Any]],
        sheet_name: str,
        statement_type: str | None,
        document_version_id: str,
    ) -> list[dict[str, Any]]:
        if not rows:
            return []
        headers = rows[0]
        cells = []
        for row_index, row in enumerate(rows[1:], start=2):
            row_label = str(row[0]).strip() if row and row[0] is not None else None
            for column_index, value in enumerate(row[1:], start=2):
                if value is None:
                    continue
                column_label = (
                    str(headers[column_index - 1]).strip()
                    if column_index - 1 < len(headers) and headers[column_index - 1] is not None
                    else None
                )
                raw_value = str(value)
                numeric_value = None
                if not raw_value.startswith("="):
                    try:
                        numeric_value = str(Decimal(raw_value.replace(",", "")))
                    except InvalidOperation:
                        pass
                year_match = re.search(r"(?<!\d)(20\d{2})(?!\d)", column_label or "")
                fiscal_year = int(year_match.group(1)) if year_match else None
                cell_reference = f"{get_column_letter(column_index)}{row_index}"
                cells.append(
                    {
                        "statement_type": statement_type,
                        "row_label": row_label,
                        "column_label": column_label,
                        "raw_value": raw_value,
                        "numeric_value": numeric_value,
                        "unit": None,
                        "currency": None,
                        "scale": None,
                        "period_start": f"{fiscal_year}-01-01" if fiscal_year else None,
                        "period_end": f"{fiscal_year}-12-31" if fiscal_year else None,
                        "fiscal_year": fiscal_year,
                        "consolidation_scope": None,
                        "audited": None,
                        "restated": None,
                        "cell_reference": cell_reference,
                        "evidence_id": stable_uuid(
                            [document_version_id, sheet_name, cell_reference]
                        ),
                    }
                )
        return cells


class XbrlParser(Parser):
    name = "lxml-xbrl"
    version = "0.1.0"
    media_types = {"application/xml", "text/xml", "application/xbrl+xml"}

    def parse(
        self,
        path: Path,
        *,
        document_id: str,
        document_version_id: str,
        document_type: DocumentType,
        artifact_id: str,
    ) -> ParsedDocument:
        root = etree.parse(str(path)).getroot()
        facts: list[dict[str, Any]] = []
        for node in root.iter():
            if len(node) or not (node.text and node.text.strip()):
                continue
            facts.append(
                {
                    "concept": etree.QName(node).localname,
                    "value": node.text.strip(),
                    "context_ref": node.get("contextRef"),
                    "unit_ref": node.get("unitRef"),
                    "decimals": node.get("decimals"),
                }
            )
        text = "\n".join(f"{fact['concept']}: {fact['value']}" for fact in facts)
        element = make_element(
            document_version_id=document_version_id,
            element_type=ElementType.STRUCTURED_RECORD,
            reading_order=0,
            text=text,
            parser_name=self.name,
            parser_version=self.version,
            structured_payload={"facts": facts},
        )
        return ParsedDocument(
            document_id=document_id,
            document_version_id=document_version_id,
            document_type=document_type,
            metadata={"artifact_id": artifact_id},
            elements=[element],
            parse_quality={"fact_count": float(len(facts))},
        )
